#!/usr/bin/env python3
"""Microsoft 365 導入提案書 Excel生成スクリプト"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ====== カラーパレット ======
COL_NAVY      = "1F3864"   # 紺（見出し背景）
COL_BLUE      = "2E75B6"   # 青（サブ見出し背景）
COL_LIGHTBLUE = "DEEAF1"   # 薄青（行ストライプ）
COL_GREEN     = "375623"   # 緑（安心・強調）
COL_LIGHTGREEN= "E2EFDA"   # 薄緑（行ストライプ）
COL_ORANGE    = "C55A11"   # オレンジ（警告・注意）
COL_LIGHTORANGE="FCE4D6"   # 薄オレンジ（行ストライプ）
COL_YELLOW    = "FFF2CC"   # 薄黄（ハイライト）
COL_WHITE     = "FFFFFF"
COL_GRAY      = "F2F2F2"

def border(style="thin", color="AAAAAA"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    th = Side(style="medium", color="1F3864")
    tn = Side(style="thin", color="AAAAAA")
    return Border(left=th, right=th, top=th, bottom=th)

def h1(ws, row, col, text, colspan=1, rowspan=1,
        bg=COL_NAVY, fg=COL_WHITE, size=14, bold=True, center=True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Meiryo UI", size=size, bold=bold, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True
    )
    cell.border = border("medium", COL_NAVY)
    if colspan > 1 or rowspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row + rowspan - 1, end_column=col + colspan - 1
        )
    return cell

def h2(ws, row, col, text, colspan=1, rowspan=1,
        bg=COL_BLUE, fg=COL_WHITE, size=11, bold=True, center=False):
    return h1(ws, row, col, text, colspan, rowspan, bg, fg, size, bold, center)

def cell(ws, row, col, text, colspan=1, rowspan=1,
         bg=None, fg="000000", size=10, bold=False,
         center=False, wrap=True, indent=0):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Meiryo UI", size=size, bold=bold, color=fg)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap, indent=indent
    )
    c.border = border()
    if colspan > 1 or rowspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row + rowspan - 1, end_column=col + colspan - 1
        )
    return c

def num_cell(ws, row, col, value, fmt="¥#,##0", bg=None, fg="000000", bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Meiryo UI", size=10, bold=bold, color=fg)
    c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border()
    return c

def set_col_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def set_row_height(ws, heights):
    for row, h in heights.items():
        ws.row_dimensions[row].height = h

# ========================================================
# Sheet 1: Office環境の提案
# ========================================================
def build_sheet1(ws):
    ws.title = "①Office環境の提案"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COL_NAVY

    set_col_width(ws, [2, 22, 30, 30, 18, 2])
    set_row_height(ws, {
        1: 8, 2: 50, 3: 30, 4: 28, 5: 15,
        6: 48, 7: 48, 8: 48,
        9: 15, 10: 28, 11: 15,
        12: 48, 13: 48, 14: 48,
        15: 15, 16: 32, 17: 15
    })

    # ── タイトル ──────────────────────────
    h1(ws, 2, 2,
       "Microsoft 365 導入提案書\n① Office環境とクラウド保存の仕組み",
       colspan=4, rowspan=1, size=16)

    # ── サブタイトル ──────────────────────
    h2(ws, 3, 2,
       "● 基本方針：既存PCのOfficeを活かしながら、保存先をOneDriveに移行する",
       colspan=4, size=10)

    # ── 現状と提案 見出し ─────────────────
    h2(ws, 4, 2, "項　目", bg=COL_NAVY, center=True)
    h2(ws, 4, 3, "現　状（NAS環境）", bg="7F7F7F", center=True)
    h2(ws, 4, 4, "提　案（M365環境）", bg=COL_BLUE, center=True)
    h2(ws, 4, 5, "効　果", bg=COL_GREEN, center=True)

    rows = [
        ("保存場所",
         "社内NAS（シノロジー）\n→ 過去にデータ消失トラブル発生",
         "OneDrive for Business\n（1TB／人）\nMicrosoft管理の冗長サーバーに自動保存",
         "データ消失リスク\nほぼゼロ"),
        ("バックアップ",
         "手動バックアップが必要\n担当者が忘れると危険",
         "保存のたびに自動でクラウド同期\n過去バージョンも30日間保持",
         "完全自動バックアップ\n履歴復元も可能"),
        ("共同編集",
         "ファイルを都度メール添付\n上書き・バージョン混在が発生",
         "複数人がリアルタイムで\n同じファイルを編集\n変更者・変更箇所が一目で判明",
         "業務効率が大幅向上\nミス防止"),
    ]

    for i, (item, current, proposal, effect) in enumerate(rows):
        r = 6 + i
        bg_a = COL_LIGHTBLUE if i % 2 == 0 else COL_WHITE
        bg_p = COL_LIGHTBLUE if i % 2 == 0 else COL_WHITE
        cell(ws, r, 2, item, bold=True, bg=COL_GRAY, center=True)
        cell(ws, r, 3, current, bg=bg_a)
        cell(ws, r, 4, proposal, bg=bg_p)
        cell(ws, r, 5, effect, bg=COL_LIGHTGREEN, bold=True, center=True, fg=COL_GREEN)

    # ── 権限管理セクション ─────────────────
    h2(ws, 10, 2, "● 厳格な権限管理（重要）", colspan=4, bg=COL_NAVY, size=10)

    perm_rows = [
        ("一般職員",
         "自分のフォルダのみ閲覧・編集\n他者フォルダは存在すら見えない",
         "情報漏えいリスクを排除"),
        ("管理職",
         "担当フォルダ＋共有フォルダにアクセス\n（権限付与された範囲のみ）",
         "必要最低限の権限で\nセキュリティを維持"),
        ("理事長",
         "全フォルダへのアクセス権\n専用フォルダは理事長のみ\n存在すら他者には見えない設定",
         "機密情報の完全保護"),
    ]

    h2(ws, 11, 2, "役　職", bg=COL_NAVY, center=True, size=10)
    h2(ws, 11, 3, "アクセス権限", bg=COL_NAVY, center=True, size=10)
    h2(ws, 11, 4, "セキュリティ効果", bg=COL_NAVY, center=True, size=10)
    h2(ws, 11, 5, "備　考", bg=COL_NAVY, center=True, size=10)

    for i, (role, access, effect) in enumerate(perm_rows):
        r = 12 + i
        bg = COL_LIGHTBLUE if i % 2 == 0 else COL_WHITE
        cell(ws, r, 2, role, bold=True, bg=COL_GRAY, center=True)
        cell(ws, r, 3, access, bg=bg)
        cell(ws, r, 4, effect, bg=bg, bold=True)
        cell(ws, r, 5, "ID＋パスワード認証\n（多要素認証も設定可）", bg=bg, fg="555555")

    # ── フッター ──────────────────────────
    cell(ws, 17, 2,
         "※ Microsoft 365 の権限設定は、初期設定費用（300,000円）に含まれます。",
         colspan=4, bg=COL_YELLOW, fg=COL_ORANGE, bold=True, size=9)


# ========================================================
# Sheet 2: 2週間の安心検証ロードマップ
# ========================================================
def build_sheet2(ws):
    ws.title = "②検証ロードマップ"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COL_BLUE

    set_col_width(ws, [2, 16, 18, 32, 26, 18, 2])
    set_row_height(ws, {
        1: 8, 2: 50, 3: 30, 4: 28,
        5: 60, 6: 60, 7: 60,
        8: 15, 9: 32, 10: 15
    })

    h1(ws, 2, 2,
       "Microsoft 365 導入提案書\n② 2週間「安心検証」ロードマップ",
       colspan=5, rowspan=1, size=16)

    h2(ws, 3, 2,
       "● いきなり全社導入せず、まず2週間の試行で「操作感」と「セキュリティ」を確認します",
       colspan=5, size=10)

    headers = ["フェーズ", "期　間", "実施内容", "確認ポイント", "担当"]
    bg_h = COL_NAVY
    for i, h in enumerate(headers):
        h2(ws, 4, 2 + i, h, bg=bg_h, center=True)

    phases = [
        ("フェーズ１\n検証環境構築",
         "1〜3日目",
         "・Microsoft 365テナントを作成\n"
         "・理事長アカウントを最初に作成\n"
         "・OneDriveの権限グループを設計\n"
         "・既存PCにOneDriveをインストール",
         "・ログインできるか\n・フォルダが正しく見えるか\n・理事長フォルダの秘匿性を確認",
         "コンサルタント\n＋IT担当者",
         COL_LIGHTBLUE),
        ("フェーズ２\n実データ検証",
         "4〜10日目",
         "・実際の業務データをOneDriveに移行\n"
         "・複数人でのリアルタイム共同編集テスト\n"
         "・権限設定の動作確認\n"
         "（他の職員に見えないことを確認）\n"
         "・既存OfficeアプリとOneDriveの連携確認",
         "・共同編集がスムーズか\n・権限違反アクセスが\n  ブロックされるか\n・保存速度・安定性の確認",
         "コンサルタント\nサポート付き",
         COL_LIGHTGREEN),
        ("フェーズ３\n結果確認・判断",
         "11〜14日目",
         "・検証レポートの作成・提出\n"
         "・操作感・セキュリティの評価\n"
         "・全6拠点への展開計画を確定\n"
         "・本導入のスケジュール決定",
         "・理事長様の承認\n・懸念点の最終確認\n・全拠点展開のGO/STOP判断",
         "コンサルタント\n＋理事長",
         COL_LIGHTORANGE),
    ]

    for i, (phase, period, content, check, person, bg) in enumerate(phases):
        r = 5 + i
        cell(ws, r, 2, phase, bold=True, bg=COL_BLUE, fg=COL_WHITE, center=True)
        cell(ws, r, 3, period, bold=True, bg=bg, center=True)
        cell(ws, r, 4, content, bg=bg)
        cell(ws, r, 5, check, bg=bg)
        cell(ws, r, 6, person, bg=bg, center=True)

    cell(ws, 9, 2,
         "✔ 2週間の検証で「問題なし」と確認できた場合のみ、全6拠点への本導入を進めます。"
         "　途中で中止しても追加費用は発生しません。",
         colspan=5, bg=COL_YELLOW, fg=COL_ORANGE, bold=True, size=9)


# ========================================================
# Sheet 3: 御見積（ハイブリッド・コスト削減プラン）
# ========================================================
def build_sheet3(ws):
    ws.title = "③御見積"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COL_GREEN

    set_col_width(ws, [2, 26, 16, 14, 14, 16, 2])
    set_row_height(ws, {
        1: 8, 2: 50, 3: 30, 4: 28,
        5: 30, 6: 30, 7: 30,
        8: 15, 9: 28, 10: 30,
        11: 15, 12: 28,
        13: 30, 14: 30,
        15: 15, 16: 50, 17: 15
    })

    h1(ws, 2, 2,
       "Microsoft 365 導入提案書\n③ 御見積（ハイブリッド・コスト削減プラン）",
       colspan=5, rowspan=1, size=16)

    h2(ws, 3, 2,
       "● 役割に応じてライセンスを使い分け、月額コストを最小化します",
       colspan=5, size=10)

    # ── ライセンス料 ───────────────────────
    h2(ws, 4, 2, "【ライセンス料（月額）】", colspan=5, bg=COL_NAVY, size=11)

    lic_headers = ["プラン", "主な用途・対象者", "単価（月額）", "想定人数", "月額小計"]
    for i, h in enumerate(lic_headers):
        h2(ws, 5, 2 + i, h, bg=COL_BLUE, center=True)

    lic_rows = [
        ("Microsoft 365\nBusiness Standard",
         "編集メイン層\n（事務・管理職など、Officeを\nよく使うスタッフ）",
         1870, 6, 11220,
         COL_LIGHTBLUE),
        ("Microsoft 365\nBusiness Basic",
         "閲覧メイン層\n（確認・参照がメインの\nスタッフ・パート等）",
         900, 6, 5400,
         COL_WHITE),
    ]

    for i, (plan, usage, unit, num, sub, bg) in enumerate(lic_rows):
        r = 6 + i
        cell(ws, r, 2, plan, bold=True, bg=bg)
        cell(ws, r, 3, usage, bg=bg)
        num_cell(ws, r, 4, unit, bg=bg)
        num_cell(ws, r, 5, num, fmt="#,##0 人", bg=bg)
        num_cell(ws, r, 6, sub, bg=bg)

    # 月額合計行
    h2(ws, 8, 2, "月額ライセンス料　合計", colspan=3, bg=COL_NAVY, center=True)
    num_cell(ws, 8, 5, "（目安）", fmt="@", bg=COL_NAVY)
    c = ws.cell(row=8, column=5, value="（規模により変動）")
    c.font = Font(name="Meiryo UI", size=9, color=COL_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COL_NAVY)
    c.border = border("medium", COL_NAVY)
    num_cell(ws, 8, 6, "約15,000〜20,000", fmt="¥#,##0 円／月", bg=COL_YELLOW, bold=True, fg=COL_ORANGE)
    c2 = ws.cell(row=8, column=6, value="約 15,000〜20,000 円／月")
    c2.font = Font(name="Meiryo UI", size=11, bold=True, color=COL_ORANGE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill = PatternFill("solid", fgColor=COL_YELLOW)
    c2.border = border("medium", COL_ORANGE)

    # ── 初期設定費 ───────────────────────
    h2(ws, 9, 2, "【初期設定費（一括・税別）】", colspan=5, bg=COL_NAVY, size=11)

    init_headers = ["内　容", "詳　細", "備　考", "", "金　額"]
    for i, h in enumerate(init_headers):
        h2(ws, 10, 2 + i, h, bg=COL_BLUE, center=True)

    init_rows = [
        ("6拠点 権限設計・構築",
         "フォルダ構成設計\n理事長専用フォルダの秘匿設定\n役職別アクセス権限の設定",
         "拠点数：6\n設計工数込み",
         COL_LIGHTBLUE),
        ("2週間 検証サポート",
         "オンサイト／リモートで検証を伴走\n操作説明・トラブル対応\n検証レポートの作成・提出",
         "検証フェーズ全期間\nサポート含む",
         COL_WHITE),
        ("スピード導入費用",
         "2週間での迅速な環境構築\n既存データの移行支援\nユーザー向け操作マニュアル作成",
         "緊急対応・\n優先対応費用込み",
         COL_LIGHTBLUE),
    ]

    for i, (name, detail, note, bg) in enumerate(init_rows):
        r = 11 + i
        cell(ws, r, 2, name, bold=True, bg=bg)
        cell(ws, r, 3, detail, bg=bg)
        cell(ws, r, 4, note, bg=bg, center=True)
        cell(ws, r, 5, "", colspan=2, bg=bg)

    ws.merge_cells(start_row=11, start_column=5, end_row=13, end_column=6)
    c = ws.cell(row=11, column=5, value="300,000 円（税別）")
    c.font = Font(name="Meiryo UI", size=14, bold=True, color=COL_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=COL_YELLOW)
    c.border = border("medium", COL_NAVY)

    # ── 年間コスト比較 ────────────────────
    h2(ws, 14, 2, "【年間コスト試算】", colspan=5, bg=COL_NAVY, size=11)

    cell(ws, 15, 2, "項　目", bold=True, bg=COL_GRAY, center=True)
    cell(ws, 15, 3, "月　額", bold=True, bg=COL_GRAY, center=True)
    cell(ws, 15, 4, "年　額", bold=True, bg=COL_GRAY, center=True)
    cell(ws, 15, 5, "備　考", bold=True, bg=COL_GRAY, center=True, colspan=2)

    cell(ws, 16, 2, "ライセンス料", bold=True, bg=COL_LIGHTBLUE)
    c = ws.cell(row=16, column=3, value="約 16,620 円")
    c.font = Font(name="Meiryo UI", size=10)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.fill = PatternFill("solid", fgColor=COL_LIGHTBLUE)
    c.border = border()
    c2 = ws.cell(row=16, column=4, value="約 199,440 円")
    c2.font = Font(name="Meiryo UI", size=10)
    c2.alignment = Alignment(horizontal="right", vertical="center")
    c2.fill = PatternFill("solid", fgColor=COL_LIGHTBLUE)
    c2.border = border()
    cell(ws, 16, 5, "Standard 6名 ＋ Basic 6名の場合", bg=COL_LIGHTBLUE, colspan=2)

    cell(ws, 17, 2, "初期設定費（1回のみ）", bold=True, bg=COL_WHITE)
    c3 = ws.cell(row=17, column=3, value="－")
    c3.font = Font(name="Meiryo UI", size=10)
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.fill = PatternFill("solid", fgColor=COL_WHITE)
    c3.border = border()
    c4 = ws.cell(row=17, column=4, value="300,000 円")
    c4.font = Font(name="Meiryo UI", size=10, bold=True, color=COL_ORANGE)
    c4.alignment = Alignment(horizontal="right", vertical="center")
    c4.fill = PatternFill("solid", fgColor=COL_WHITE)
    c4.border = border()
    cell(ws, 17, 5, "初年度のみ発生", bg=COL_WHITE, colspan=2)

    cell(ws, 18, 2, "初年度　合計概算", bold=True, bg=COL_YELLOW)
    c5 = ws.cell(row=18, column=3, value="－")
    c5.font = Font(name="Meiryo UI", size=10)
    c5.alignment = Alignment(horizontal="center", vertical="center")
    c5.fill = PatternFill("solid", fgColor=COL_YELLOW)
    c5.border = border("medium", COL_ORANGE)
    c6 = ws.cell(row=18, column=4, value="約 499,440 円")
    c6.font = Font(name="Meiryo UI", size=13, bold=True, color=COL_ORANGE)
    c6.alignment = Alignment(horizontal="right", vertical="center")
    c6.fill = PatternFill("solid", fgColor=COL_YELLOW)
    c6.border = border("medium", COL_ORANGE)
    cell(ws, 18, 5,
         "2年目以降はライセンス料のみ\n（約20万円／年）",
         bold=True, bg=COL_YELLOW, fg=COL_GREEN, colspan=2)

    set_row_height(ws, {14: 28, 15: 28, 16: 32, 17: 32, 18: 36})

    # ── フッター ──────────────────────────
    set_row_height(ws, {19: 15, 20: 36})
    cell(ws, 20, 2,
         "※ 本見積は概算です。最終人数・プラン構成により変動します。"
         "　　詳細は別途ご相談ください。　　有効期限：提示日より30日間",
         colspan=5, bg=COL_GRAY, fg="555555", size=9)


# ========================================================
# メイン
# ========================================================
def main():
    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()

    build_sheet1(ws1)
    build_sheet2(ws2)
    build_sheet3(ws3)

    out = "/home/user/pikaichi.APP/Microsoft365_導入提案書.xlsx"
    wb.save(out)
    print(f"✅ 保存完了: {out}")

if __name__ == "__main__":
    main()
